import os
from xlutils.copy import copy
import xlrd as ExcelRead


def write_append(file_name):
    values = ["Ann", "woman", 22, "UK"]

    r_xls = ExcelRead.open_workbook(file_name)
    r_sheet = r_xls.sheet_by_index(0)
    rows = r_sheet.nrows
    w_xls = copy(r_xls)
    sheet_write = w_xls.get_sheet(0)

    for i in range(0, len(values)):
        sheet_write.write(rows, i, values[i])

    w_xls.save(file_name + '.out' + os.path.splitext(file_name)[-1]);


if __name__ == "__main__":
    write_append("./test_append.xls")

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

def write_excel():
    ex=load_workbook(filename=r'D:\PycharmProjects\hupitest\data\case.xlsx')
    print('open excel success!')
    ws = ex.get_sheet_by_name("case")
    print('open sheet1 success!')
    ws.cell(row=4, column=2).value = 'hupi2222'
    print('write values success!')
    ex.save(filename='D:\PycharmProjects\hupitest\data\case.xlsx')
    print('save success!')


if __name__ == '__main__':
    write_excel()

import xlrd
import os
from xlutils.copy import copy
from xlwt import easyxf
import xlwt

#写入excel
def excel_write():
    #excel=r'D:\PycharmProjects\hupitest\data\TestBalanceQueryCase.xls'
    rb = xlrd.open_workbook('D:\PycharmProjects\hupitest\data\TestBalanceQueryCase.xls')
    wb = copy(rb)
    print(wb)
    sheet = wb.get_sheet(0)
    print(sheet)
    sheet.write(2, 4, 'hupitest')
    #os.remove(excel)
    print('write success')
    wb.save('D:\PycharmProjects\hupitest\data\new.xls')
    print('save excel success！')
