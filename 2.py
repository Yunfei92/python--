from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
def append_data():
    ex = load_workbook(filename=r'D:\\人民网排行榜.xlsx)
    print('open excel success!')
    ws = ex.get_sheet_by_name("人民网排行榜")
    print('open sheet1 success!')
    ws.cell(row=50, column=2).value = 'hupi2222'
    print('write values success!')
    ex.save(filename=r'D:\\人民网排行榜.xlsx)
    print('save success!')
    import xlwt;
    import xlrd;
    # import xlutils;
    from xlutils.copy import copy;

    styleBoldRed = xlwt.easyxf('font: color-index red, bold on');
    headerStyle = styleBoldRed;
    wb = xlwt.Workbook();
    ws = wb.add_sheet(gConst['xls']['sheetName']);
    ws.write(0, 0, "Header", headerStyle);
    ws.write(0, 1, "CatalogNumber", headerStyle);
    ws.write(0, 2, "PartNumber", headerStyle);
    wb.save(gConst['xls']['fileName']);

    # open existed xls file
    # newWb = xlutils.copy(gConst['xls']['fileName']);
    # newWb = copy(gConst['xls']['fileName']);
    oldWb = xlrd.open_workbook(gConst['xls']['fileName'], formatting_info=True);
    print
    oldWb;  # <xlrd.book.Book object at 0x000000000315C940>
    newWb = copy(oldWb);
    print
    newWb;  # <xlwt.Workbook.Workbook object at 0x000000000315F470>
    newWs = newWb.get_sheet(0);
    newWs.write(1, 0, "value1");
    newWs.write(1, 1, "value2");
    newWs.write(1, 2, "value3");
    print
    "write new values ok";
    newWb.save(gConst['xls']['fileName']);
    print
    "save with same name ok";